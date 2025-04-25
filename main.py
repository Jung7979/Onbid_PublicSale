import requests
import pandas as pd
import xml.etree.ElementTree as ET
from datetime import datetime
import time
from tqdm import tqdm
import os
import sys
import math
from multiprocessing import Pool, cpu_count
from functools import partial
import numpy as np
import openpyxl  # openpyxl 모듈 추가
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.styles import Alignment
from datetime import datetime, timedelta
from dotenv import load_dotenv

# .env 파일 로드
load_dotenv()

# 서비스 키 가져오기
SERVICE_KEY = os.getenv('API_KEY_KAMCO_Decoding')
if not SERVICE_KEY:
    print(".env 파일에 API_KEY_KAMCO_Decoding를 설정해주세요.")
    sys.exit(1)

class KamcoAuctionService:
    def __init__(self, service_key):
        self.base_url = "http://openapi.onbid.co.kr/openapi/services/UtlinsttPblsalThingInquireSvc"
        self.service_key = service_key
        self.backup_folder = os.path.join(os.getcwd(), "backup")
        self.data_folder = os.path.join(self.backup_folder, "data")
        
        # 폴더 생성
        for folder in [self.backup_folder, self.data_folder]:
            if not os.path.exists(folder):
                os.makedirs(folder)
                print(f"폴더 생성: {folder}")

    def get_total_count(self, disposal_method='0001'):
        """
        전체 데이터 개수 조회
        """
        endpoint = f"{self.base_url}/getPublicSaleObject"
        params = {
            'serviceKey': self.service_key,
            'numOfRows': 1,
            'pageNo': 1,
            'DPSL_MTD_CD': disposal_method
        }

        try:
            response = requests.get(endpoint, params=params)
            response.raise_for_status()
            
            root = ET.fromstring(response.content)
            
            # 결과 코드 확인
            result_code = root.find('.//resultCode').text
            if result_code != '00':
                result_msg = root.find('.//resultMsg').text
                raise Exception(f"API Error: {result_code} - {result_msg}")
                
            total_count = int(root.find('.//totalCount').text)
            return total_count
        
        except requests.exceptions.RequestException as e:
            raise Exception(f"Request failed: {str(e)}")
        except ET.ParseError as e:
            raise Exception(f"XML parsing failed: {str(e)}")
        except Exception as e:
            raise Exception(f"Error occurred: {str(e)}")
        
    def get_total_count(self, disposal_method='0001'):
        """
        전체 데이터 개수 조회
        """
        # 일주일 전 날짜 계산
        week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y%m%d')
        
        endpoint = f"{self.base_url}/getPublicSaleObject"
        params = {
            'serviceKey': self.service_key,
            'numOfRows': 1,
            'pageNo': 1,
            'DPSL_MTD_CD': disposal_method,
            'PBCT_BEGN_DTM': week_ago  # 일주일 전 날짜 추가
        }

        try:
            response = requests.get(endpoint, params=params)
            response.raise_for_status()
            
            root = ET.fromstring(response.content)
            
            # 결과 코드 확인
            result_code = root.find('.//resultCode').text
            if result_code != '00':
                result_msg = root.find('.//resultMsg').text
                raise Exception(f"API Error: {result_code} - {result_msg}")
                
            total_count = int(root.find('.//totalCount').text)
            return total_count
        
        except requests.exceptions.RequestException as e:
            raise Exception(f"Request failed: {str(e)}")
        except ET.ParseError as e:
            raise Exception(f"XML parsing failed: {str(e)}")
        except Exception as e:
            raise Exception(f"Error occurred: {str(e)}")

    def get_auction_items(self, num_of_rows=100, page_no=1, disposal_method='0001'):
        """
        공매물건 목록 조회
        """
        # 일주일 전 날짜 계산
        week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y%m%d')

        endpoint = f"{self.base_url}/getPublicSaleObject"
        params = {
            'serviceKey': self.service_key,
            'numOfRows': num_of_rows,
            'pageNo': page_no,
            'DPSL_MTD_CD': disposal_method,
            'PBCT_BEGN_DTM': week_ago  # 일주일 전 날짜 추가

        }

        try:
            response = requests.get(endpoint, params=params)
            response.raise_for_status()
            
            root = ET.fromstring(response.content)
            
            # 결과 코드 확인
            result_code = root.find('.//resultCode').text
            if result_code != '00':
                result_msg = root.find('.//resultMsg').text
                raise Exception(f"API Error: {result_code} - {result_msg}")

            items = []
            for item in root.findall('.//item'):
                item_data = self.get_item_data(item)
                items.append(item_data)
                
            return items

        except requests.exceptions.RequestException as e:
            raise Exception(f"Request failed: {str(e)}")
        except ET.ParseError as e:
            raise Exception(f"XML parsing failed: {str(e)}")
        except Exception as e:
            raise Exception(f"Error occurred: {str(e)}")

    def get_item_data(self, item):
        """
        XML 항목에서 모든 데이터 추출하여 한글 필드명으로 변환
        """
        # 영문-한글 필드 매핑
        field_mapping = {
            'RNUM': '순번',
            'PLNM_NO': '공고번호',
            'PBCT_NO': '공매번호',
            'PBCT_CDTN_NO': '공매조건번호',
            'CLTR_NO': '물건번호',
            'CLTR_HSTR_NO': '물건이력번호',
            'SCRN_GRP_CD': '화면그룹코드',
            'CTGR_FULL_NM': '용도명',
            'BID_MNMT_NO': '입찰번호',
            'CLTR_NM': '물건명',
            'CLTR_MNMT_NO': '물건관리번호',
            'LDNM_ADRS': '물건소재지(지번)',
            'NMRD_ADRS': '물건소재지(도로명)',
            'LDNM_PNU': '지번PNU',
            'DPSL_MTD_CD': '처분방식코드',
            'DPSL_MTD_NM': '처분방식코드명',
            'BID_MTD_NM': '입찰방식명',
            'MIN_BID_PRC': '최저입찰가',
            'APSL_ASES_AVG_AMT': '감정가',
            'FEE_RATE': '최저입찰가율',
            'PBCT_BEGN_DTM': '입찰시작일시',
            'PBCT_CLS_DTM': '입찰마감일시',
            'PBCT_CLTR_STAT_NM': '물건상태',
            'USCBD_CNT': '유찰횟수',
            'IQRY_CNT': '조회수',
            'GOODS_NM': '물건상세정보',
            'MANF': '제조사',
            'MDL': '모델',
            'NRGT': '연월식',
            'GRBX': '변속기',
            'ENDPC': '배기량',
            'VHCL_MLGE': '주행거리',
            'FUEL': '연료',
            'SCRT_NM': '법인명',
            'TPBZ': '업종',
            'ITM_NM': '종목명',
            'MMB_RGT_NM': '회원권명',
            'CLTR_IMG_FILE': '물건 이미지'
        }

        # XML 데이터 추출 및 한글 필드명으로 변환
        data = {}
        for eng_field, kor_field in field_mapping.items():
            value = item.find(eng_field)
            data[kor_field] = value.text if value is not None else ''
        
        return data

    def fetch_page_data(self, page_info):
        """
        단일 페이지 데이터 수집 (병렬 처리용)
        """
        page_no, disposal_method, items_per_page = page_info
        
        for attempt in range(3):  # 재시도 횟수
            try:
                time.sleep(0.5)  # API 호출 간격
                items = self.get_auction_items(
                    num_of_rows=items_per_page,
                    page_no=page_no,
                    disposal_method=disposal_method
                )
                return items
            except Exception as e:
                if attempt == 2:  # 마지막 시도
                    print(f"\n페이지 {page_no} 처리 실패: {str(e)}")
                    return []
                time.sleep(2)  # 재시도 전 대기
        return []

    def save_data_to_excel(self, items, filename, is_backup=False):
        """
        데이터를 엑셀 파일로 저장
        """
        if not items:
            print("저장할 데이터가 없습니다.")
            return
            
        try:
            # 절대 경로로 변환
            abs_filename = os.path.abspath(filename)
            print(f"파일 저장 시도: {abs_filename}")
            
            df = pd.DataFrame(items)

            # 헤더 순서 정의
            columns_order = [
                '순번',
                '물건관리번호',
                '용도명',
                '물건명',
                '물건소재지(지번)',
                '지번PNU',
                '물건소재지(도로명)',
                '입찰방식명',
                '감정가',
                '최저입찰가',
                '최저입찰가율',
                '입찰시작일시',
                '입찰마감일시',
                '물건상태',
                '유찰횟수',
                '조회수',
                '물건상세정보',
                '공고번호',
                '공매번호',
                '공매조건번호',
                '물건번호',
                '물건이력번호',
                '화면그룹코드',
                '입찰번호',
                '처분방식코드',
                '처분방식코드명',
                '제조사',
                '모델',
                '연월식',
                '변속기',
                '배기량',
                '주행거리',
                '연료',
                '법인명',
                '업종',
                '종목명',
                '회원권명',
                '물건 이미지'
            ]
            
            # 존재하는 컬럼만 선택하고 순서대로 정렬
            existing_columns = [col for col in columns_order if col in df.columns]
            df = df[existing_columns]
            
            # 엑셀 파일 생성
            with pd.ExcelWriter(abs_filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='공매물건목록')
                
                # 워크시트 가져오기
                worksheet = writer.sheets['공매물건목록']
                
                # 스타일 설정
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 열 너비 자동 조정
                for column in worksheet.columns:
                    max_length = 0
                    column = list(column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

                # 하이퍼링크 추가
                for row in range(2, worksheet.max_row + 1):  # 2부터 시작 (헤더 제외)
                    cltr_hstr_no = worksheet.cell(row=row, column=existing_columns.index('물건이력번호') + 1).value
                    cltr_no = worksheet.cell(row=row, column=existing_columns.index('물건번호') + 1).value
                    plnm_no = worksheet.cell(row=row, column=existing_columns.index('공고번호') + 1).value
                    pbct_no = worksheet.cell(row=row, column=existing_columns.index('공매번호') + 1).value
                    scrn_grp_cd = worksheet.cell(row=row, column=existing_columns.index('화면그룹코드') + 1).value
                    pbct_cdtn_no = worksheet.cell(row=row, column=existing_columns.index('공매조건번호') + 1).value

                    if all([cltr_hstr_no, cltr_no, plnm_no, pbct_no, scrn_grp_cd, pbct_cdtn_no]):
                        url = f"https://www.onbid.co.kr/op/cta/cltrdtl/collateralDetailMoveableAssetsDetail.do?cltrHstrNo={cltr_hstr_no}&cltrNo={cltr_no}&plnmNo={plnm_no}&pbctNo={pbct_no}&scrnGrpCd={scrn_grp_cd}&pbctCdtnNo={pbct_cdtn_no}"
                        cell = worksheet.cell(row=row, column=existing_columns.index('물건관리번호') + 1)
                        cell.hyperlink = url
                        cell.font = Font(color="0000FF", underline="single")
            
            print(f"파일 저장 완료: {abs_filename}")
            
        except Exception as e:
            print(f"파일 저장 중 오류 발생: {str(e)}")
            raise

    def process_chunk(self, chunk_data, chunk_number, total_chunks):
        """
        데이터 청크 처리 및 저장
        """
        if not chunk_data:
            print(f"청크 {chunk_number}에 데이터가 없습니다.")
            return
        
        try:
            current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
            chunk_filename = os.path.join(
                self.data_folder,
                f"kamco_auction_chunk_{chunk_number}_of_{total_chunks}_{current_time}.xlsx"
            )
            
            print(f"청크 데이터 저장 시도: {chunk_filename}")
            self.save_data_to_excel(chunk_data, chunk_filename, is_backup=True)
            print(f"청크 데이터 저장 완료: {chunk_filename} ({len(chunk_data):,}건)")
            
        except Exception as e:
            print(f"청크 저장 중 오류 발생: {str(e)}")
            raise

    def merge_chunk_files(self):
        """
        청크 파일들을 하나로 병합
        """
        try:
            all_data = []
            chunk_files = [f for f in os.listdir(self.data_folder) if f.startswith("kamco_auction_chunk_")]
            
            print(f"\n병합할 청크 파일 수: {len(chunk_files)}")
            
            for file in chunk_files:
                file_path = os.path.join(self.data_folder, file)
                print(f"파일 처리 중: {file_path}")
                
                try:
                    df = pd.read_excel(file_path)
                    all_data.append(df)
                    print(f"파일 처리 완료: {file_path}")
                except Exception as e:
                    print(f"파일 처리 중 오류 발생: {file_path} - {str(e)}")
            
            if all_data:
                merged_data = pd.concat(all_data, ignore_index=True)
                return merged_data.to_dict('records')
            return []
            
        except Exception as e:
            print(f"청크 파일 병합 중 오류 발생: {str(e)}")
            raise
    
    def get_all_items(self, disposal_method='0001', items_per_page=100, chunk_size=1000):
        """
        전체 공매물건 데이터 수집 (최적화된 버전)
        """
        try:
            total_count = self.get_total_count(disposal_method)
            print(f"\n전체 데이터 개수: {total_count:,}개")
            
            total_pages = (total_count + items_per_page - 1) // items_per_page
            
            # 페이지 정보 생성
            page_infos = [(page, disposal_method, items_per_page) 
                         for page in range(1, total_pages + 1)]
            
            # CPU 코어 수 제한 (4개만 사용)
            num_processes = min(4, cpu_count())
            print(f"\n{num_processes}개의 프로세스로 병렬 처리 시작")
            
            all_items = []
            current_chunk = []
            chunk_count = 0

            # 오류 복구를 위한 재시도 횟수
            max_retries = 3
            
            with Pool(processes=num_processes) as pool:
                with tqdm(total=total_pages, desc="데이터 수집 중") as pbar:
                    try:
                        # imap 사용 (순차적 처리, 더 안정적)
                        for items in pool.imap(self.fetch_page_data, page_infos):
                            if items:
                                all_items.extend(items)
                                current_chunk.extend(items)
                                
                                # chunk_size에 도달하면 청크 저장
                                if len(current_chunk) >= chunk_size:
                                    try:
                                        chunk_count += 1
                                        self.process_chunk(
                                            current_chunk, 
                                            chunk_count, 
                                            math.ceil(total_count / chunk_size)
                                        )
                                        # 성공적으로 저장된 후에만 청크 초기화
                                        current_chunk = []
                                    except Exception as e:
                                        print(f"\n청크 저장 중 오류 발생: {str(e)}")
                            
                            pbar.update(1)  # 수정된 부분: pbar.update(1) 위치 이동
                            # 중간 진행상황 저장
                            if len(all_items) % (chunk_size * 5) == 0:
                                try:
                                    backup_filename = os.path.join(
                                        self.backup_folder,
                                        f"kamco_auction_backup_{len(all_items)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                                    )
                                    self.save_data_to_excel(all_items, backup_filename, is_backup=True)
                                    print(f"\n중간 백업 완료: {backup_filename} (총 {len(all_items):,}건)")
                                except Exception as e:
                                    print(f"\n중간 백업 중 오류 발생: {str(e)}")
                        
                        pbar.update(1)
                        pbar.set_postfix({'수집': f'{len(all_items):,}건'})
                
                    except KeyboardInterrupt:
                        print("\n사용자에 의해 중단됨. 지금까지 수집된 데이터 저장 중...")
                        # 중단 시점까지의 데이터 저장
                        if all_items:
                            try:
                                interrupt_filename = os.path.join(
                                    self.backup_folder,
                                    f"kamco_auction_interrupted_{len(all_items)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                                )
                                self.save_data_to_excel(all_items, interrupt_filename, is_backup=True)
                                print(f"\n중단 시점 데이터 저장 완료: {interrupt_filename}")
                            except Exception as e:
                                print(f"\n중단 데이터 저장 중 오류 발생: {str(e)}")
                        raise
                    
                    except Exception as e:
                        print(f"\n데이터 수집 중 오류 발생: {str(e)}")
                        if all_items:
                            try:
                                error_filename = os.path.join(
                                    self.backup_folder,
                                    f"kamco_auction_error_{len(all_items)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                                )
                                self.save_data_to_excel(all_items, error_filename, is_backup=True)
                                print(f"\n오류 발생 시점 데이터 저장 완료: {error_filename}")
                            except Exception as save_error:
                                print(f"\n오류 데이터 저장 실패: {str(save_error)}")
                        raise
        
            # 남은 청크 처리
            if current_chunk:
                try:
                    chunk_count += 1
                    self.process_chunk(
                        current_chunk, 
                        chunk_count, 
                        math.ceil(total_count / chunk_size)
                    )
                except Exception as e:
                    print(f"\n최종 청크 저장 중 오류 발생: {str(e)}")
            
            # 최종 파일 저장
            try:
                final_filename = os.path.join(
                    self.backup_folder,
                    f"kamco_auction_full_{len(all_items)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                )
                self.save_data_to_excel(all_items, final_filename, is_backup=True)
                print(f"\n최종 데이터 저장 완료: {final_filename} (총 {len(all_items):,}건)")
            except Exception as e:
                print(f"\n최종 데이터 저장 중 오류 발생: {str(e)}")
            
            print(f"\n수집된 전체 데이터 개수: {len(all_items):,}개")
            return all_items
        
        except Exception as e:
            print(f"\n치명적 오류 발생: {str(e)}")
            raise

    def fetch_page_data(self, page_info):
        """
        단일 페이지 데이터 수집 (개선된 버전)
        """
        page_no, disposal_method, items_per_page = page_info
        
        for attempt in range(3):
            try:
                # API 호출 간격 증가 (2초)
                time.sleep(2)
                
                items = self.get_auction_items(
                    num_of_rows=items_per_page,
                    page_no=page_no,
                    disposal_method=disposal_method
                )
                return items
            except Exception as e:
                if attempt == 2:  # 마지막 시도
                    print(f"\n페이지 {page_no} 처리 실패: {str(e)}")
                    return []
                time.sleep(5)  # 재시도 전 대기 시간 증가
        return []

def main():
    try:
        print("이용기관 공고 목록 조회 서비스 시작")
        
        service = KamcoAuctionService(SERVICE_KEY)
        
        # chunk_size를 조정하여 메모리 사용량과 성능 최적화
        items = service.get_all_items(
            items_per_page=100,  # API 호출당 데이터 수
            chunk_size=1000      # 청크당 데이터 수
        )
        
        print("\n프로그램 종료")
        
    except Exception as e:
        print(f"\nError: {str(e)}")

if __name__ == "__main__":
    main()